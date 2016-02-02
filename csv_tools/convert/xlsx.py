#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
*******************
Excel (2007 and up)
*******************

This takes data from a given xlsx Sheet and converts it to a
CSV. It is using the wonderful OpenPyXL library to extract
the data from the Sheet.

"""

import datetime

from openpyxl.reader.excel import load_workbook
import six

from csv_tools import CSVToolsWriter
from csv_tools.typeinference import NULL_TIME

__author__ = "Daniel Melichar"
__copyright__ = "Copyright 2015"
__credits__ = ["Daniel Melichar"]
__license__ = "GPL"
__version__ = "1.0"
__maintainer__ = "Daniel Melichar"
__email__ = "dmelichar@student.tgm.ac.at"
__status__ = "Deployed"


def normalize_datetime(dt):
    """
    Normalizes a datetime cell
    :param dt: datetime element
    :return:
    """

    if dt.microsecond == 0:
        return dt

    ms = dt.microsecond

    if ms < 1000:
        return dt.replace(microsecond=0)
    elif ms > 999000:
        return dt.replace(microsecond=0) + datetime.timedelta(seconds=1)

    return dt


def has_date_elements(cell):
    """
    Try to use formatting to determine if a cell contains only time info.
    See: http://office.microsoft.com/en-us/excel-help/number-format-codes-HP005198679.aspx

    :param cell: The cell to look at
    """
    if 'd' in cell.number_format or \
                    'y' in cell.number_format:
        return True

    return False


def xlsx(f, output=None, **kwargs):
    """
    Convert an Excel .xlsx file to csv.
    Note: Unlike other convertor's, this one allows output columns to contain mixed data types.
    Blank headers are also possible.

    :param f: Excel File to convert
    :type f: File
    :param output: CSV
    :type output: File
    """
    streaming = True if output else False

    if not streaming:
        output = six.StringIO()

    writer = CSVToolsWriter(output)

    book = load_workbook(f, use_iterators=True, data_only=True)

    if 'sheet' in kwargs:
        sheet = book.get_sheet_by_name(kwargs['sheet'])
    else:
        sheet = book.get_active_sheet()

    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            writer.writerow([c.value for c in row])
            continue

        out_row = []

        for c in row:
            value = c.value

            if value.__class__ is datetime.datetime:
                # Handle default XLSX date as 00:00 time
                if value.date() == datetime.date(1904, 1, 1) and not has_date_elements(c):
                    value = value.time()

                    value = normalize_datetime(value)
                elif value.time() == NULL_TIME:
                    value = value.date()
                else:
                    value = normalize_datetime(value)
            elif value.__class__ is float:
                if value % 1 == 0:
                    value = int(value)

            if value.__class__ in (datetime.datetime, datetime.date, datetime.time):
                value = value.isoformat()

            out_row.append(value)

        writer.writerow(out_row)

    if not streaming:
        data = output.getvalue()
        return data

    # Return empty string when streaming
    return ''
